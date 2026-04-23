from typing import Optional
from datetime import datetime, timezone
import asyncio
import io
import httpx
import openpyxl
import xlrd
from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from pydantic import BaseModel, field_validator
from sqlalchemy import case
from sqlalchemy.orm import Session
from slowapi import Limiter
from slowapi.util import get_remote_address
from auth import get_current_user
from audit import log_action
from database import get_db
import models

router = APIRouter(prefix="/game-requests", tags=["game-requests"])
limiter = Limiter(key_func=get_remote_address)

VALID_STATUSES = {"pending", "top", "done"}
PAGE_SIZE = 50
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB cap on Salla contact uploads


def _get_db():
    yield from get_db()


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class GameRequestSubmit(BaseModel):
    game: str
    order_number: Optional[str] = None

    @field_validator("game")
    @classmethod
    def game_not_empty(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("game is required")
        if len(v) > 200:
            raise ValueError("game name must be 200 characters or fewer")
        return v

    @field_validator("order_number")
    @classmethod
    def order_number_max(cls, v: Optional[str]) -> Optional[str]:
        if v is not None:
            v = v.strip() or None
            if v and len(v) > 50:
                raise ValueError("order_number must be 50 characters or fewer")
        return v


class GameRequestUpdate(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None

    @field_validator("status")
    @classmethod
    def status_valid(cls, v: Optional[str]) -> Optional[str]:
        if v is not None and v not in VALID_STATUSES:
            raise ValueError(f"status must be one of: {', '.join(sorted(VALID_STATUSES))}")
        return v


class SteamLinkBody(BaseModel):
    app_id: str
    name: str
    url: str
    price_uah: Optional[float] = None
    price_sar: Optional[float] = None
    discount: Optional[int] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _serialize(r: models.GameRequest) -> dict:
    return {
        "id": r.id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "game_name": r.game_name,
        "order_number": r.order_number,
        "status": r.status,
        "notes": r.notes,
        "steam_app_id": r.steam_app_id,
        "steam_name": r.steam_name,
        "steam_url": r.steam_url,
        "steam_price_uah": r.steam_price_uah,
        "steam_price_sar": r.steam_price_sar,
        "steam_discount": r.steam_discount,
        "deleted_at": r.deleted_at.isoformat() if r.deleted_at else None,
    }


# ── Public endpoint (no auth) ─────────────────────────────────────────────────

@router.post("")
@limiter.limit("5/minute")
def submit_game_request(
    body: GameRequestSubmit,
    request: Request,
    db: Session = Depends(_get_db),
):
    try:
        gr = models.GameRequest(
            game_name=body.game,
            order_number=body.order_number,
            status="pending",
        )
        db.add(gr)
        db.commit()
        db.refresh(gr)
    except Exception:
        db.rollback()
        raise HTTPException(status_code=400, detail="Failed to save game request")

    log_action(
        db,
        action="game_request_submitted",
        detail=body.game,
        ip=request.client.host if request.client else None,
    )

    return {"ok": True}


# ── Admin endpoints (auth required) ──────────────────────────────────────────

@router.get("")
def list_game_requests(
    status: Optional[str] = None,
    page: int = 1,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    q = db.query(models.GameRequest)
    if status == "deleted":
        q = q.filter(models.GameRequest.deleted_at.isnot(None))
    else:
        q = q.filter(models.GameRequest.deleted_at.is_(None))
        if status:
            q = q.filter(models.GameRequest.status == status)
        else:
            q = q.filter(models.GameRequest.status != "done")
    total = q.count()
    # Top items first, then by date desc
    priority = case({"top": 0}, value=models.GameRequest.status, else_=1)
    items = (
        q.order_by(priority, models.GameRequest.created_at.desc())
        .offset((page - 1) * PAGE_SIZE)
        .limit(PAGE_SIZE)
        .all()
    )
    return {
        "total": total,
        "page": page,
        "pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
        "items": [_serialize(r) for r in items],
    }


@router.get("/steam-search")
async def steam_search(
    q: str,
    current_user: models.User = Depends(get_current_user),
):
    """Search Steam Ukraine store and return up to 5 results with UAH + SAR prices."""
    # Single HTTP/2 client reused across all calls → one connection pool, no repeated TLS.
    async with httpx.AsyncClient(timeout=15.0, http2=True) as client:
        # 1. Search Steam store
        search_resp = await client.get(
            "https://store.steampowered.com/api/storesearch/",
            params={"term": q, "l": "english", "cc": "UA"},
        )
        search_resp.raise_for_status()
        search_data = search_resp.json()

        items = (search_data.get("items") or [])[:5]
        if not items:
            return []

        # 2. Fetch exchange rate UAH → SAR (open.er-api.com supports UAH)
        sar_rate: Optional[float] = None
        try:
            fx_resp = await client.get("https://open.er-api.com/v6/latest/UAH")
            fx_resp.raise_for_status()
            sar_rate = fx_resp.json().get("rates", {}).get("SAR")
        except Exception:
            sar_rate = None

        # 3. Fetch app details concurrently (reuses same client + connection pool)
        async def fetch_details(app_id: str):
            try:
                resp = await client.get(
                    "https://store.steampowered.com/api/appdetails",
                    params={"appids": app_id, "cc": "UA", "l": "english"},
                )
                resp.raise_for_status()
                return app_id, resp.json()
            except Exception:
                return app_id, None

        app_ids = [str(item["id"]) for item in items]
        detail_results = await asyncio.gather(*[fetch_details(aid) for aid in app_ids])
        details_map = {aid: data for aid, data in detail_results}

    # 4. Build response
    results = []
    for item in items:
        app_id = str(item["id"])
        name = item.get("name", "")
        url = f"https://store.steampowered.com/app/{app_id}"

        detail_data = details_map.get(app_id)
        is_free = False
        not_available = False
        price_uah: Optional[float] = None
        price_sar: Optional[float] = None
        discount_percent: int = 0

        if detail_data and detail_data.get(app_id, {}).get("success"):
            app_info = detail_data[app_id]["data"]
            if app_info.get("is_free"):
                is_free = True
                price_uah = 0.0
                price_sar = 0.0
            else:
                price_overview = app_info.get("price_overview")
                if price_overview:
                    # Steam returns price in minor units (kopecks), divide by 100
                    raw = price_overview.get("final", 0)
                    price_uah = round(raw / 100, 2)
                    if sar_rate is not None:
                        price_sar = round(price_uah * sar_rate, 2)
                    discount_percent = price_overview.get("discount_percent", 0)
                else:
                    not_available = True
        else:
            not_available = True

        results.append({
            "app_id": app_id,
            "name": name,
            "url": url,
            "price_uah": price_uah,
            "price_sar": price_sar,
            "is_free": is_free,
            "not_available": not_available,
            "discount_percent": discount_percent,
        })

    return results


@router.patch("/{request_id}/steam")
def link_steam(
    request_id: int,
    body: SteamLinkBody,
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.is_(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")

    gr.steam_app_id = body.app_id
    gr.steam_name = body.name
    gr.steam_url = body.url
    gr.steam_price_uah = body.price_uah
    gr.steam_price_sar = body.price_sar
    gr.steam_discount = body.discount

    db.commit()
    db.refresh(gr)

    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_request_steam_linked",
        resource="game_requests",
        detail=f"id={request_id} steam_app_id={body.app_id}",
        ip=request.client.host if request.client else None,
    )

    return _serialize(gr)


@router.delete("/{request_id}")
def delete_game_request(
    request_id: int,
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.is_(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")

    gr.deleted_at = datetime.now(timezone.utc)
    db.commit()

    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_request_deleted",
        resource="game_requests",
        detail=f"id={request_id} game_name={gr.game_name}",
        ip=request.client.host if request.client else None,
    )

    return {"ok": True}


@router.post("/{request_id}/restore")
def restore_game_request(
    request_id: int,
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.isnot(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")

    gr.deleted_at = None
    db.commit()

    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_request_restored",
        resource="game_requests",
        detail=f"id={request_id} game_name={gr.game_name}",
        ip=request.client.host if request.client else None,
    )

    return _serialize(gr)


@router.delete("/{request_id}/permanent")
def permanent_delete_game_request(
    request_id: int,
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.isnot(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")

    game_name = gr.game_name
    db.delete(gr)
    db.commit()

    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_request_permanently_deleted",
        resource="game_requests",
        detail=f"id={request_id} game_name={game_name}",
        ip=request.client.host if request.client else None,
    )

    return {"ok": True}


@router.post("/refresh-prices")
async def refresh_all_prices(
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Re-fetch Ukraine Steam prices for all linked game requests."""
    linked = (
        db.query(models.GameRequest)
        .filter(models.GameRequest.steam_app_id.isnot(None))
        .all()
    )
    if not linked:
        return {"updated": 0}

    async with httpx.AsyncClient(timeout=15.0) as client:
        # Fetch exchange rate once
        sar_rate: Optional[float] = None
        try:
            fx_resp = await client.get("https://open.er-api.com/v6/latest/UAH")
            fx_resp.raise_for_status()
            sar_rate = fx_resp.json().get("rates", {}).get("SAR")
        except Exception:
            sar_rate = None

        async def fetch_price(app_id: str):
            try:
                resp = await client.get(
                    "https://store.steampowered.com/api/appdetails",
                    params={"appids": app_id, "cc": "UA", "l": "english"},
                )
                resp.raise_for_status()
                return app_id, resp.json()
            except Exception:
                return app_id, None

        app_ids = [gr.steam_app_id for gr in linked]
        results = await asyncio.gather(*[fetch_price(aid) for aid in app_ids])
        details_map = {aid: data for aid, data in results}

    updated = 0
    for gr in linked:
        detail_data = details_map.get(gr.steam_app_id)
        if not detail_data or not detail_data.get(gr.steam_app_id, {}).get("success"):
            continue
        app_info = detail_data[gr.steam_app_id]["data"]
        if app_info.get("is_free"):
            gr.steam_price_uah = 0.0
            gr.steam_price_sar = 0.0
            gr.steam_discount = 0
            updated += 1
        else:
            price_overview = app_info.get("price_overview")
            if price_overview:
                raw = price_overview.get("final", 0)
                gr.steam_price_uah = round(raw / 100, 2)
                if sar_rate is not None:
                    gr.steam_price_sar = round(gr.steam_price_uah * sar_rate, 2)
                gr.steam_discount = price_overview.get("discount_percent", 0)
                updated += 1

    db.commit()
    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_requests_prices_refreshed",
        resource="game_requests",
        detail=f"updated={updated}",
        ip=request.client.host if request.client else None,
    )
    return {"updated": updated}


@router.post("/upload-contacts")
async def upload_contacts(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Upload a Salla orders XLS/XLSX to populate order → name + phone lookup."""
    # Enforce size cap before loading into memory.
    if file.size is not None and file.size > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    rows = []

    try:
        # Try openpyxl first (handles .xlsx and .xls files that are actually xlsx)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lstrip('\ufeff').strip('"') if c.value else "" for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except Exception:
            # Fall back to xlrd for genuine old .xls files
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            for r in range(1, ws.nrows):
                rows.append({headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)})
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")


    now = datetime.now(timezone.utc)
    upserted = 0

    # Parse all rows first to build the set of order numbers
    parsed = []
    for row in rows:
        order_no = str(row.get("رقم الطلب") or "").strip().split(".")[0]
        name = str(row.get("اسم العميل") or "").strip()
        phone = str(row.get("رقم الجوال") or "").strip().lstrip("+")
        if not order_no or order_no in ("", "nan", "None"):
            continue
        parsed.append((order_no, name, phone))

    # Bulk-load all existing contacts in one query
    order_nos = {p[0] for p in parsed}
    existing_map = {
        c.order_number: c
        for c in db.query(models.SallaOrderContact)
        .filter(models.SallaOrderContact.order_number.in_(order_nos))
        .all()
    }

    for order_no, name, phone in parsed:
        existing = existing_map.get(order_no)
        if existing:
            existing.customer_name = name or existing.customer_name
            existing.phone = phone or existing.phone
            existing.uploaded_at = now
        else:
            db.add(models.SallaOrderContact(
                order_number=order_no,
                customer_name=name,
                phone=phone,
                uploaded_at=now,
            ))
        upserted += 1

    db.commit()
    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="salla_contacts_uploaded",
        resource="salla_order_contacts",
        detail=f"upserted={upserted}",
        ip=request.client.host if request.client else None,
    )
    return {"upserted": upserted}


@router.get("/{request_id}/notify-info")
def notify_info(
    request_id: int,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Look up customer name and phone from uploaded Salla contacts by order number."""
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.is_(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")
    if not gr.order_number:
        raise HTTPException(status_code=400, detail="No order number on this request")

    contact = db.query(models.SallaOrderContact).filter(
        models.SallaOrderContact.order_number == str(gr.order_number)
    ).first()

    if not contact:
        raise HTTPException(status_code=404, detail="No contact found — upload a Salla orders file first")

    return {
        "name": contact.customer_name or "العميل",
        "phone": contact.phone or "",
        "game_name": gr.game_name,
    }


@router.patch("/{request_id}")
def update_game_request(
    request_id: int,
    body: GameRequestUpdate,
    request: Request,
    db: Session = Depends(_get_db),
    current_user: models.User = Depends(get_current_user),
):
    gr = db.query(models.GameRequest).filter(
        models.GameRequest.id == request_id,
        models.GameRequest.deleted_at.is_(None),
    ).first()
    if not gr:
        raise HTTPException(status_code=404, detail="Game request not found")

    changes = []
    if body.status is not None:
        changes.append(f"status: {gr.status} → {body.status}")
        gr.status = body.status
    if body.notes is not None:
        gr.notes = body.notes
        changes.append("notes updated")

    db.commit()
    db.refresh(gr)

    log_action(
        db,
        user_id=current_user.id,
        username=current_user.username,
        action="game_request_updated",
        resource="game_requests",
        detail=f"id={request_id} " + "; ".join(changes),
        ip=request.client.host if request.client else None,
    )

    return _serialize(gr)
